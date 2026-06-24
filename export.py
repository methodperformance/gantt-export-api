"""
Vercel Serverless Function — Gantt Excel Export
Receives JSON action data, injects into template, returns .xlsx file
"""
from http.server import BaseHTTPRequestHandler
import json, datetime, re, base64, io, os
from openpyxl import load_workbook

SECTION_ROWS = {7, 11, 20, 50, 58, 65, 71}
TASK_ROWS = (
    list(range(8,  11)) +
    list(range(12, 20)) +
    list(range(21, 50)) +
    list(range(51, 58)) +
    list(range(59, 65)) +
    list(range(66, 71))
)

def get_template_b64():
    """Load template from file, return as base64"""
    template_path = os.path.join(os.path.dirname(__file__), '..', 'gantt_template.xlsx')
    with open(template_path, 'rb') as f:
        return base64.b64encode(f.read()).decode()

def generate_excel(project):
    """Inject project actions into template, return bytes"""
    template_path = os.path.join(os.path.dirname(__file__), '..', 'gantt_template.xlsx')
    wb = load_workbook(template_path)
    ws = wb['Gantt']

    actions = project.get('actions', [])

    # Clear task rows
    for r in TASK_ROWS:
        for col in [1, 2, 3, 4, 6, 8]:
            ws.cell(r, col).value = None

    # Inject data
    for i, a in enumerate(actions):
        if i >= len(TASK_ROWS): break
        r = TASK_ROWS[i]

        ws.cell(r, 1).value = '    ' + (a.get('name') or '').strip()
        ws.cell(r, 2).value = (a.get('owner') or '').strip()

        def parse_date(s):
            if not s: return None
            try: return datetime.date.fromisoformat(str(s)[:10])
            except: return None

        start = parse_date(a.get('start'))
        end   = parse_date(a.get('end'))
        if start: ws.cell(r, 3).value = start
        if end:   ws.cell(r, 4).value = end

        pct = a.get('percent', 0) or 0
        ws.cell(r, 6).value = pct / 100 if pct > 1 else pct
        ws.cell(r, 8).value = (a.get('update') or '').strip()

    # Save to bytes buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


class handler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        """Handle CORS preflight"""
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_POST(self):
        try:
            # Read request body
            length  = int(self.headers.get('Content-Length', 0))
            body    = self.rfile.read(length)
            project = json.loads(body)

            # Generate Excel
            xlsx_bytes = generate_excel(project)

            # Build filename
            plant  = re.sub(r'[^A-Za-z0-9_-]', '_', project.get('plant', 'PL'))
            leader = re.sub(r'[^A-Za-z0-9_-]', '_', project.get('leader', 'LDR'))
            name   = re.sub(r'[^A-Za-z0-9_-]', '_', project.get('name', 'Project'))
            fname  = f'{plant}_{leader}_{name}.xlsx'

            # Return file
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
            self.send_header('Content-Length', str(len(xlsx_bytes)))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(xlsx_bytes)

        except Exception as e:
            error = json.dumps({'error': str(e)}).encode()
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Length', str(len(error)))
            self.end_headers()
            self.wfile.write(error)

    def do_GET(self):
        """Health check"""
        msg = json.dumps({'status': 'ok', 'service': 'Gantt Export API'}).encode()
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Content-Length', str(len(msg)))
        self.end_headers()
        self.wfile.write(msg)

    def log_message(self, format, *args):
        pass  # Suppress default logging
